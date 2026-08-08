# ==============================================================================
# OTOMATISASI ANALISIS RUAS JALAN LAPANGAN DARI FILE KMZ
# Versi: Master 12 Kolom + Fixed Scope NameError & Auto Timestamp
# ==============================================================================

import os
import sys
import zipfile
import tempfile
import xml.etree.ElementTree as ET
import urllib.request
import json
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
from shapely.geometry import LineString
from pyproj import Transformer
import openpyxl
import threading 
import math
import subprocess
import logging
import shutil
import time
import datetime

# Setup logging
logging.basicConfig(filename='app_debug.log', level=logging.DEBUG, format='%(asctime)s %(levelname)s:%(message)s')
_api_cache = {}

def decimal_to_dms(lat, lon):
    """Konversi koordinat desimal ke format DMS presisi."""
    if lat is None or lon is None or pd.isna(lat) or pd.isna(lon): return ""
    lat_dir, lat_abs = ('N', abs(lat)) if lat >= 0 else ('S', abs(lat))
    lat_d = int(lat_abs)
    lat_m = int((lat_abs - lat_d) * 60)
    lat_s = round((lat_abs - lat_d - lat_m/60) * 3600, 2)
    if lat_s >= 60.0: lat_s, lat_m = lat_s - 60.0, lat_m + 1
    if lat_m >= 60: lat_m, lat_d = lat_m - 60, lat_d + 1
        
    lon_dir, lon_abs = ('E', abs(lon)) if lon >= 0 else ('W', abs(lon))
    lon_d = int(lon_abs)
    lon_m = int((lon_abs - lon_d) * 60)
    lon_s = round((lon_abs - lon_d - lon_m/60) * 3600, 2)
    if lon_s >= 60.0: lon_s, lon_m = lon_s - 60.0, lon_m + 1
    if lon_m >= 60: lon_m, lon_d = lon_m - 60, lon_d + 1
        
    return f'{lat_d}°{lat_m}\'{lat_s}"{lat_dir} {lon_d}°{lon_m}\'{lon_s}"{lon_dir}'

def get_utm_epsg(lon, lat):
    zone_number = int((lon + 180) / 6) + 1
    return f"EPSG:326{zone_number:02d}" if lat >= 0 else f"EPSG:327{zone_number:02d}"

def calculate_deflection_angle(p1, p2, p3):
    v1_x, v1_y = p2[0] - p1[0], p2[1] - p1[1]
    v2_x, v2_y = p3[0] - p2[0], p3[1] - p2[1]
    b1 = math.degrees(math.atan2(v1_y, v1_x))
    b2 = math.degrees(math.atan2(v2_y, v2_x))
    diff = abs(b1 - b2)
    return 360 - diff if diff > 180 else diff

def extract_kml_from_kmz(kmz_path):
    temp_dir = tempfile.mkdtemp()
    try:
        with zipfile.ZipFile(kmz_path, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)
        for root, _, files in os.walk(temp_dir):
            for file in files:
                if file.endswith('.kml'): return os.path.join(root, file), temp_dir
        raise FileNotFoundError("Tidak ditemukan file KML di KMZ.")
    except Exception as e:
        shutil.rmtree(temp_dir, ignore_errors=True)
        raise e

def parse_kml_all_linestrings(kml_path):
    tree = ET.parse(kml_path)
    all_coords = []
    clean_tag = lambda tag: tag.split('}')[-1] if '}' in tag else tag
    for elem in tree.getroot().iter():
        if clean_tag(elem.tag) == 'LineString':
            for child in elem.iter():
                if clean_tag(child.tag) == 'coordinates' and child.text:
                    for pt in child.text.strip().replace('\n', ' ').split():
                        parts = pt.split(',')
                        if len(parts) >= 2:
                            try:
                                lon, lat = float(parts[0]), float(parts[1])
                                if not all_coords or all_coords[-1] != (lon, lat):
                                    all_coords.append((lon, lat))
                            except ValueError: continue
                    break
    return all_coords

def get_accurate_road_info(lat, lon):
    coord_key = (round(lat, 5), round(lon, 5))
    if coord_key in _api_cache: return _api_cache[coord_key]

    road_name = "Jalan Belum Terdata (OSM)"
    authority = "Kota"
    instansi = "Dinas PUPR Kota/Kabupaten"

    try:
        url = f"https://photon.komoot.io/reverse?lon={lon}&lat={lat}"
        req = urllib.request.Request(url, headers={'User-Agent': 'GIS_Drafter_App/13.0'})
        with urllib.request.urlopen(req, timeout=5) as response:
            data = json.loads(response.read().decode())
            if data.get('features'):
                props = data['features'][0].get('properties', {})
                highway_type = props.get('highway', '')
                
                road = props.get('street') or props.get('name') or props.get('highway')
                if road: road_name = str(road).title()
                else:
                    if props.get('village'): road_name = f"Jalan Akses {str(props.get('village')).title()}"
                    elif props.get('county'): road_name = f"Kawasan {str(props.get('county')).title()}"
                
                name_lower = road_name.lower()
                
                if any(kw in name_lower for kw in ['lintas', 'nasional', 'raya ']) or highway_type in ['trunk', 'primary', 'motorway']:
                    authority, instansi = "Provinsi", "Dinas PUPR Provinsi"
                elif "belum terdata" in name_lower or "akses" in name_lower or "gang " in name_lower or highway_type in ['track', 'path', 'unclassified']:
                    authority, instansi = "Desa", "Pemerintah Desa"
                else:
                    authority, instansi = "Kota", "Dinas PUPR Kota/Kabupaten"
            
            result = (road_name, authority, instansi)
            _api_cache[coord_key] = result
            time.sleep(0.02)
            return result
    except Exception: return (road_name, authority, instansi)

def process_single_kmz(kmz_path, spk, ring_id, area_name, update_status_func):
    kml_path, temp_dir = extract_kml_from_kmz(kmz_path)
    try:
        points = parse_kml_all_linestrings(kml_path)
        if len(points) < 2: return []

        transformer = Transformer.from_crs("EPSG:4326", get_utm_epsg(points[0][0], points[0][1]), always_xy=True)
        segments_data = []
        seg_start_idx, last_api_idx = 0, 0
        
        current_road, current_auth, current_instansi = get_accurate_road_info(points[0][1], points[0][0])
        
        i = 1
        while i < len(points):
            lon, lat = points[i]
            cur_utm = transformer.transform(lon, lat)
            last_utm = transformer.transform(points[last_api_idx][0], points[last_api_idx][1])
            
            is_sharp = calculate_deflection_angle(points[i-1], points[i], points[i+1]) > 30 if i < len(points)-1 else False
            dist = math.hypot(cur_utm[0] - last_utm[0], cur_utm[1] - last_utm[1])
            
            if (i == len(points) - 1) or is_sharp or dist >= 50:
                if update_status_func and i % 20 == 0: update_status_func(f"Memindai vertex {i}/{len(points)} pada {ring_id}...")
                chk_road, chk_auth, chk_inst = get_accurate_road_info(lat, lon)
                
                if chk_road and chk_road != current_road:
                    cut_idx = i
                    for j in range(last_api_idx + 1, i):
                        m_road, m_auth, m_inst = get_accurate_road_info(points[j][1], points[j][0])
                        if m_road and m_road != current_road:
                            cut_idx, chk_road, chk_auth, chk_inst = j, m_road, m_auth, m_inst
                            break
                    
                    seg_pts = points[seg_start_idx : cut_idx + 1]
                    if len(seg_pts) > 1:
                        line_utm = LineString([transformer.transform(pt[0], pt[1]) for pt in seg_pts])
                        segments_data.append([
                            spk, ring_id, "", area_name, current_auth, current_instansi,
                            current_road, round(line_utm.length, 2), "",
                            decimal_to_dms(seg_pts[0][1], seg_pts[0][0]), decimal_to_dms(seg_pts[-1][1], seg_pts[-1][0]),
                            ""
                        ])
                    
                    current_road, current_auth, current_instansi = chk_road, chk_auth, chk_inst
                    seg_start_idx = cut_idx
                    last_api_idx = cut_idx
                    i = cut_idx + 1
                    continue
                last_api_idx = i
            i += 1

        if seg_start_idx < len(points) - 1:
            seg_pts = points[seg_start_idx : len(points)]
            if len(seg_pts) > 1:
                line_utm = LineString([transformer.transform(pt[0], pt[1]) for pt in seg_pts])
                segments_data.append([
                    spk, ring_id, "", area_name, current_auth, current_instansi,
                    current_road, round(line_utm.length, 2), "",
                    decimal_to_dms(seg_pts[0][1], seg_pts[0][0]), decimal_to_dms(seg_pts[-1][1], seg_pts[-1][0]),
                    ""
                ])
        return segments_data
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Otomatisasi Rekap Ruas Jalan KMZ - Master Builder")
        self.root.geometry("620x380")
        self.root.resizable(False, False)
        self.kmz_files = []
        
        frame = ttk.Frame(root, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Otomatisasi Hitung & Excel Master 12 Kolom", font=('Helvetica', 11, 'bold')).pack(anchor='w', pady=(0, 10))
        
        self.btn_select = ttk.Button(frame, text="📁 Pilih File KMZ Input (Bisa Banyak File)", command=self.select_files)
        self.btn_select.pack(fill=tk.X, pady=5)
        
        self.lbl_file = ttk.Label(frame, text="File terpilih: Belum ada file", foreground="gray", wraplength=580)
        self.lbl_file.pack(anchor='w', pady=2)
        
        self.btn_process = ttk.Button(frame, text="⚡ Proses Semua Menjadi 1 File Excel Master", command=self.run_process_threaded, state=tk.DISABLED)
        self.btn_process.pack(fill=tk.X, pady=(15, 5))
        
        self.progress = ttk.Progressbar(frame, orient="horizontal", mode="determinate")
        self.progress.pack(fill=tk.X, pady=5)
        
        self.lbl_status = ttk.Label(frame, text="", foreground="green")
        self.lbl_status.pack(anchor='w', pady=5)

    def select_files(self):
        filenames = filedialog.askopenfilenames(title="Pilih File KMZ", filetypes=[("KMZ Files", "*.kmz")])
        if filenames:
            self.kmz_files = list(filenames)
            self.lbl_file.config(text=f"Total file terpilih: {len(self.kmz_files)} file KMZ", foreground="black")
            self.btn_process.config(state=tk.NORMAL)
            self.lbl_status.config(text="")

    def run_process_threaded(self):
        self.btn_process.config(state=tk.DISABLED)
        self.btn_select.config(state=tk.DISABLED)
        threading.Thread(target=self.run_process, daemon=True).start()

    def run_process(self):
        total_files = len(self.kmz_files)
        all_master_data = []
        excel_path = "" # Inisialisasi variabel di scope utama agar tidak NameError
        
        out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
        os.makedirs(out_dir, exist_ok=True)

        for index, filepath in enumerate(self.kmz_files, 1):
            basename = os.path.splitext(os.path.basename(filepath))[0]
            
            if "-" in basename:
                parts = basename.split("-", 1)
                full_code = parts[0].strip()
                area_name = parts[1].strip()
            else:
                full_code = basename.strip()
                area_name = ""
                
            if len(full_code) >= 4 and full_code[:4].isdigit():
                spk = full_code[:4]
                ring_id = full_code[4:]
            else:
                spk = full_code[:5]
                ring_id = full_code

            self.root.after(0, lambda idx=index: self.progress.config(value=(idx-1)/total_files * 100))
            self.root.after(0, lambda msg=f"[{index}/{total_files}] Memproses {spk} - {ring_id}...": self.lbl_status.config(text=msg, foreground="blue"))
            
            try:
                segments = process_single_kmz(filepath, spk, ring_id, area_name, 
                                              lambda msg: self.root.after(0, lambda: self.lbl_status.config(text=msg)))
                all_master_data.extend(segments)
            except Exception as e:
                logging.exception(f"Gagal memproses {basename}")

        # Pembuatan File Master Excel
        if all_master_data:
            columns = ["SPK", "Ring ID", "Destination", "Area", "Authority Ruas Jalan", "Instansi", 
                       "Nama Ruas Jalan Implementasi", "Panjang Ruas Jalan (Meter)", "Status Cable", 
                       "Titik Koordinat Awal", "Titik Koordinat Akhir", "Status Survey"]
            
            df_master = pd.DataFrame(all_master_data, columns=columns)
            time_stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_path = os.path.join(out_dir, f"Rekap_Master_Ruas_Jalan_{time_stamp}.xlsx")
            
            try:
                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    df_master.to_excel(writer, index=False, sheet_name='Master Data')
                    ws = writer.sheets['Master Data']
                    ws.views.sheetView[0].showGridLines = True
                    
                    hdr_fill = openpyxl.styles.PatternFill(start_color="003366", end_color="003366", fill_type="solid")
                    hdr_font = openpyxl.styles.Font(name="Arial", size=10, bold=True, color="FFFFFF")
                    border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin', color='D9D9D9'),
                                                    right=openpyxl.styles.Side(style='thin', color='D9D9D9'),
                                                    top=openpyxl.styles.Side(style='thin', color='D9D9D9'),
                                                    bottom=openpyxl.styles.Side(style='thin', color='D9D9D9'))
                    
                    for col_num in range(1, 13):
                        cell = ws.cell(row=1, column=col_num)
                        cell.fill, cell.font, cell.border = hdr_fill, hdr_font, border
                        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
                    
                    for row in range(2, len(df_master) + 2):
                        for col in range(1, 13):
                            cell = ws.cell(row=row, column=col)
                            cell.border = border
                            if col == 8:
                                cell.number_format = '#,##0.00'
                                cell.alignment = openpyxl.styles.Alignment(horizontal="right")
                            elif col in [1, 2, 5, 9, 10, 11, 12]:
                                cell.alignment = openpyxl.styles.Alignment(horizontal="center")
                            else:
                                cell.alignment = openpyxl.styles.Alignment(horizontal="left")

                    widths = {'A':10, 'B':16, 'C':15, 'D':22, 'E':20, 'F':25, 'G':35, 'H':25, 'I':15, 'J':30, 'K':30, 'L':18}
                    for col_letter, width in widths.items(): ws.column_dimensions[col_letter].width = width

                self.root.after(0, lambda: self.progress.config(value=100))
                self.root.after(0, lambda: self.lbl_status.config(text="Selesai! 1 File Master berhasil dibuat.", foreground="green"))
                
                try: os.startfile(out_dir) if sys.platform == "win32" else subprocess.run(["open" if sys.platform == "darwin" else "xdg-open", out_dir])
                except Exception: pass
                
                # Menggunakan variabel tersimpan di scope aman
                final_path = excel_path
                self.root.after(0, lambda: messagebox.showinfo("Sukses Sempurna", f"Berhasil menggabungkan {total_files} file KMZ.\n\nTersimpan di:\n{final_path}"))
            except Exception as err:
                logging.exception("Error saat menyimpan Excel Master")
                self.root.after(0, lambda: messagebox.showerror("Error Simpan File", f"Gagal menyimpan file Excel:\n{err}"))
                
        self.root.after(0, lambda: self.btn_process.config(state=tk.NORMAL))
        self.root.after(0, lambda: self.btn_select.config(state=tk.NORMAL))

if __name__ == "__main__":
    app = App(tk.Tk())
    app.root.mainloop()